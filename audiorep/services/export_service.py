"""
ExportService — Exportación de la biblioteca de pistas.

Formatos soportados:
    XLSX — dos hojas: "Biblioteca" + "Estadísticas" (requiere openpyxl).
    PDF  — dos secciones con el mismo contenido (requiere fpdf2).
    CSV  — biblioteca o estadísticas por separado (stdlib csv).

Métodos principales:
    export_xlsx(tracks, stats, filepath)          — biblioteca + estadísticas combinadas
    export_pdf(tracks, stats, filepath)           — ídem en PDF
    export_csv(tracks, filepath)                  — solo biblioteca

    export_library_xlsx(tracks, filepath)         — solo hoja Biblioteca
    export_stats_xlsx(stats, filepath)            — solo hoja Estadísticas
    export_library_pdf(tracks, filepath)          — solo sección pistas
    export_stats_pdf(stats, filepath)             — solo sección estadísticas
    export_stats_csv(stats, filepath)             — estadísticas en CSV (Sección,Indicador,Valor)
"""
from __future__ import annotations

import csv
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from audiorep.domain.track import Track
    from audiorep.services.stats_service import LibraryStats


# ---------------------------------------------------------------------------
# Utilidades internas
# ---------------------------------------------------------------------------

def _ms_to_str(ms: int) -> str:
    if not ms:
        return "0:00"
    total_s = ms // 1000
    m, s = divmod(total_s, 60)
    return f"{m}:{s:02d}"


def _stars(rating: int) -> str:
    """Versión Unicode para XLSX (soporta ★☆)."""
    if rating <= 0:
        return "Sin rating"
    return "★" * rating + "☆" * (5 - rating)


def _stars_text(rating: int) -> str:
    """Versión ASCII para PDF (Helvetica no soporta ★☆)."""
    if rating <= 0:
        return "Sin rating"
    return f"{rating}/5"


# ---------------------------------------------------------------------------
# ExportService
# ---------------------------------------------------------------------------

class ExportService:
    """Exporta pistas y estadísticas a diferentes formatos de archivo."""

    # ------------------------------------------------------------------
    # Helpers privados XLSX
    # ------------------------------------------------------------------

    @staticmethod
    def _write_library_sheet(ws: object, tracks: list[Track]) -> None:
        """Rellena una hoja de openpyxl con la tabla de pistas."""
        from openpyxl.styles import Alignment, Font, PatternFill

        hdr_font  = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill  = PatternFill("solid", fgColor="2D2D2D")
        alt_fill  = PatternFill("solid", fgColor="F2F2F2")
        data_font = Font(color="1A1A1A", size=11)
        center    = Alignment(horizontal="center")
        left      = Alignment(horizontal="left")

        headers    = ["#", "Título", "Artista", "Álbum", "Año", "Género", "Duración", "Formato"]
        col_widths = [5,    44,       28,         32,      7,     22,       12,          10]

        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=h)  # type: ignore[union-attr]
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = center
            ws.column_dimensions[cell.column_letter].width = w  # type: ignore[union-attr]
        ws.row_dimensions[1].height = 16  # type: ignore[union-attr]

        for row_idx, t in enumerate(tracks, 2):
            is_alt = row_idx % 2 == 0
            vals = [
                row_idx - 1,
                t.title       or "",
                t.artist_name or "",
                t.album_title or "",
                t.year        or "",
                t.genre       or "",
                _ms_to_str(t.duration_ms),
                t.format.value if t.format else "",
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)  # type: ignore[union-attr]
                cell.font      = data_font
                cell.alignment = center if col in (1, 5, 7, 8) else left
                if is_alt:
                    cell.fill = alt_fill

        ws.freeze_panes = "A2"  # type: ignore[union-attr]

    @staticmethod
    def _write_stats_sheet(ws: object, stats: LibraryStats) -> None:
        """Rellena una hoja de openpyxl con las estadísticas y gráficos al costado."""
        from openpyxl.styles import Font, PatternFill
        from openpyxl.chart import BarChart, PieChart, Reference  # type: ignore[import]

        ws.column_dimensions["A"].width = 35  # type: ignore[union-attr]
        ws.column_dimensions["B"].width = 22  # type: ignore[union-attr]
        ws.column_dimensions["C"].width = 22  # type: ignore[union-attr]

        section_font = Font(bold=True, size=13, color="1A1A1A")
        label_font   = Font(bold=True, color="FFFFFF", size=11)
        value_font   = Font(color="1A1A1A", size=11)
        hdr_fill     = PatternFill("solid", fgColor="2D2D2D")
        alt_fill     = PatternFill("solid", fgColor="F2F2F2")

        row = 1
        # chart_defs: (anchor_row, header_row, data_start, data_end, chart_type, title)
        chart_defs: list[tuple[int, int, int, int, str, str]] = []

        def _sec(title: str) -> int:
            nonlocal row
            ws.cell(row=row, column=1, value=title).font = section_font  # type: ignore[union-attr]
            r = row; row += 1; return r

        def _hdr(*cols: str) -> int:
            nonlocal row
            for i, col in enumerate(cols, 1):
                cell = ws.cell(row=row, column=i, value=col)  # type: ignore[union-attr]
                cell.font = label_font; cell.fill = hdr_fill
            r = row; row += 1; return r

        def _row(*vals: object) -> None:
            nonlocal row
            is_alt = row % 2 == 0
            for i, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=i, value=val)  # type: ignore[union-attr]
                cell.font = value_font
                if is_alt: cell.fill = alt_fill
            row += 1

        def _section_charted(
            title: str, col_h: tuple[str, str], data: list[tuple[object, object]],
            chart_type: str,
        ) -> None:
            nonlocal row
            anchor = _sec(title)
            hr     = _hdr(*col_h)
            ds     = row
            for item in data:
                _row(*item)
            de = row - 1
            if de >= ds:
                chart_defs.append((anchor, hr, ds, de, chart_type, title))
            row += 1

        hours = stats.total_duration_ms / 3_600_000
        _sec("Resumen general")
        _hdr("Indicador", "Valor")
        _row("Total pistas",               stats.total_tracks)
        _row("Total artistas",             stats.total_artists)
        _row("Total álbumes",              stats.total_albums)
        _row("Horas de música",            f"{hours:.1f} h")
        _row("Nacionalidades de artistas", stats.total_countries)
        row += 1

        if stats.genre_counts:
            _section_charted("Géneros", ("Género", "Pistas"),
                             sorted(stats.genre_counts.items(), key=lambda x: x[1], reverse=True),
                             "bar_col")
        if stats.decade_counts:
            _section_charted("Décadas", ("Década", "Pistas"),
                             sorted(stats.decade_counts.items()),
                             "bar_col")
        if stats.format_counts:
            _section_charted("Formatos", ("Formato", "Pistas"),
                             sorted(stats.format_counts.items(), key=lambda x: x[1], reverse=True),
                             "pie")
        if stats.top_artists:
            _section_charted("Top 10 artistas por cantidad de pistas", ("Artista", "Pistas"),
                             stats.top_artists, "bar_horiz")
        if stats.album_type_counts:
            _section_charted("Tipo de álbum", ("Tipo", "Álbumes"),
                             sorted(stats.album_type_counts.items(), key=lambda x: x[1], reverse=True),
                             "bar_col")
        if stats.artist_country_counts:
            _section_charted("País de origen de artistas", ("País", "Artistas"),
                             sorted(stats.artist_country_counts.items(), key=lambda x: x[1], reverse=True),
                             "bar_horiz")
        if stats.label_country_counts:
            _section_charted("País de origen de sellos", ("País", "Sellos"),
                             sorted(stats.label_country_counts.items(), key=lambda x: x[1], reverse=True),
                             "bar_horiz")

        # ── Insertar gráficos al costado de cada sección ─────────────── #
        for anchor, hr, ds, de, ct, title in chart_defs:
            if ct == "pie":
                chart: BarChart | PieChart = PieChart()
                chart.title  = title
                chart.width  = 15
                chart.height = 8
                data_ref = Reference(ws, min_col=2, min_row=ds, max_row=de)
                cats_ref = Reference(ws, min_col=1, min_row=ds, max_row=de)
                chart.add_data(data_ref)
                chart.set_categories(cats_ref)
            else:
                chart = BarChart()
                chart.type     = "bar" if ct == "bar_horiz" else "col"  # type: ignore[union-attr]
                chart.grouping = "clustered"                             # type: ignore[union-attr]
                chart.title    = title
                chart.width    = 15
                chart.height   = 8
                data_ref = Reference(ws, min_col=2, min_row=hr, max_row=de)
                cats_ref = Reference(ws, min_col=1, min_row=ds, max_row=de)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
            ws.add_chart(chart, f"E{anchor}")  # type: ignore[union-attr]

    # ------------------------------------------------------------------
    # Helpers privados PDF
    # ------------------------------------------------------------------

    @staticmethod
    def _write_library_pdf(pdf: object, tracks: list[Track], stats: LibraryStats) -> None:
        """Añade la sección de biblioteca al objeto FPDF."""
        import fpdf as fpdf_mod
        assert isinstance(pdf, fpdf_mod.FPDF)

        pdf.add_page()
        pdf.set_font("Helvetica", "B", 14)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(0, 10, "AudioRep - Biblioteca de pistas", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", size=9)
        pdf.set_text_color(80, 80, 80)
        hours = stats.total_duration_ms / 3_600_000
        pdf.cell(
            0, 5,
            f"{stats.total_tracks} pistas  |  {stats.total_artists} artistas  |  "
            f"{stats.total_albums} albums  |  {hours:.1f} horas",
            new_x="LMARGIN", new_y="NEXT",
        )
        pdf.ln(3)

        col_w   = [8, 52, 32, 36, 10, 20, 14, 12]
        headers = ["#", "Titulo", "Artista", "Album", "Ano", "Genero", "Dur.", "Fmt"]

        pdf.set_fill_color(210, 210, 210)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "B", 9)
        for w, h in zip(col_w, headers):
            pdf.cell(w, 6, h, border=1, fill=True)
        pdf.ln()

        for i, t in enumerate(tracks, 1):
            is_alt = i % 2 == 0
            pdf.set_fill_color(242, 242, 242) if is_alt else pdf.set_fill_color(255, 255, 255)
            pdf.set_text_color(30, 30, 30)
            pdf.set_font("Helvetica", size=9)
            row_vals = [
                str(i),
                (t.title or "")[:38],
                (t.artist_name or "")[:22],
                (t.album_title or "")[:24],
                str(t.year or ""),
                (t.genre or "")[:14],
                _ms_to_str(t.duration_ms),
                (t.format.value if t.format else "")[:6],
            ]
            for w, val in zip(col_w, row_vals):
                pdf.cell(w, 5, val, border="B", fill=True)
            pdf.ln()

    @staticmethod
    def _write_stats_pdf(pdf: object, stats: LibraryStats) -> None:
        """Añade la sección de estadísticas con layout 2 columnas (tabla + gráfico)."""
        import fpdf as fpdf_mod
        assert isinstance(pdf, fpdf_mod.FPDF)

        # ── Constantes de layout ──────────────────────────────────────── #
        LEFT_X:  float = 10.0
        LEFT_W:  float = 90.0
        RIGHT_X: float = 105.0
        RIGHT_W: float = 85.0
        HDR_H:   float = 6.0
        DATA_H:  float = 5.0
        TW: list[float] = [62.0, 28.0]   # anchos columnas tabla (suma = LEFT_W)

        pdf.add_page()
        pdf.set_font("Helvetica", "B", 14)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(0, 10, "AudioRep - Estadisticas de la biblioteca",
                 new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        def _check_space(needed: float) -> None:
            available = pdf.h - pdf.get_y() - pdf.b_margin  # type: ignore[union-attr]
            if available < needed:
                pdf.add_page()  # type: ignore[union-attr]

        def _section_title(title: str) -> None:
            pdf.set_font("Helvetica", "B", 11)   # type: ignore[union-attr]
            pdf.set_text_color(40, 40, 40)        # type: ignore[union-attr]
            pdf.cell(0, 7, title, new_x="LMARGIN", new_y="NEXT")  # type: ignore[union-attr]

        def _table_left(headers: list[str], rows: list[tuple]) -> float:
            """Escribe la tabla en columna izquierda. Retorna Y final."""
            pdf.set_x(LEFT_X)                    # type: ignore[union-attr]
            pdf.set_fill_color(210, 210, 210)    # type: ignore[union-attr]
            pdf.set_text_color(0, 0, 0)          # type: ignore[union-attr]
            pdf.set_font("Helvetica", "B", 10)   # type: ignore[union-attr]
            for h, w in zip(headers, TW):
                pdf.cell(w, HDR_H, h, border=1, fill=True)  # type: ignore[union-attr]
            pdf.ln()                             # type: ignore[union-attr]
            for i, row_data in enumerate(rows):
                pdf.set_x(LEFT_X)                # type: ignore[union-attr]
                if i % 2 == 0:
                    pdf.set_fill_color(242, 242, 242)   # type: ignore[union-attr]
                else:
                    pdf.set_fill_color(255, 255, 255)   # type: ignore[union-attr]
                pdf.set_text_color(30, 30, 30)   # type: ignore[union-attr]
                pdf.set_font("Helvetica", size=9)  # type: ignore[union-attr]
                for val, w in zip(row_data, TW):
                    max_ch = max(5, int(w / 1.9))
                    pdf.cell(w, DATA_H, str(val)[:max_ch], border="B", fill=True)  # type: ignore[union-attr]
                pdf.ln()                          # type: ignore[union-attr]
            pdf.ln(2)                             # type: ignore[union-attr]
            return pdf.get_y()                    # type: ignore[union-attr]

        def _bars_right(labels: list[str], values: list[int], y0: float) -> float:
            """Dibuja barras horizontales en columna derecha. Retorna Y final."""
            if not values:
                return y0
            max_v = max(values) or 1
            lbl_w = 33.0
            val_w = 10.0
            bar_max = RIGHT_W - lbl_w - val_w
            y = y0
            for label, value in zip(labels, values):
                pdf.set_xy(RIGHT_X, y)           # type: ignore[union-attr]
                pdf.set_font("Helvetica", size=8)  # type: ignore[union-attr]
                pdf.set_text_color(30, 30, 30)   # type: ignore[union-attr]
                pdf.cell(lbl_w, DATA_H, label[:18])  # type: ignore[union-attr]
                bar_w = (value / max_v) * bar_max
                if bar_w > 0.5:
                    pdf.set_fill_color(92, 61, 159)   # type: ignore[union-attr]
                    pdf.rect(RIGHT_X + lbl_w, y + 1.0, bar_w, DATA_H - 2.0, "F")  # type: ignore[union-attr]
                pdf.set_xy(RIGHT_X + lbl_w + bar_w + 1.0, y)  # type: ignore[union-attr]
                pdf.set_font("Helvetica", size=7)  # type: ignore[union-attr]
                pdf.set_text_color(60, 60, 60)   # type: ignore[union-attr]
                pdf.cell(val_w, DATA_H, str(value))  # type: ignore[union-attr]
                y += DATA_H + 1.0
            return y + 2.0

        def _section_2col(title: str, headers: list[str], rows: list[tuple]) -> None:
            """Sección con tabla a la izquierda y gráfico de barras a la derecha."""
            if not rows:
                return
            n = min(len(rows), 20)
            est_h = 7.0 + HDR_H + n * (DATA_H + 1.0) + 5.0
            _check_space(max(est_h, 75.0))
            pdf.set_auto_page_break(False)       # type: ignore[union-attr]
            _section_title(title)
            y0 = pdf.get_y()                     # type: ignore[union-attr]
            labels = [str(r[0])[:18] for r in rows[:n]]
            int_vals: list[int] = []
            for r in rows[:n]:
                try:    int_vals.append(int(r[1]))
                except (ValueError, TypeError): int_vals.append(0)
            t_end = _table_left(headers, rows[:n])
            c_end = _bars_right(labels, int_vals, y0 + HDR_H)
            pdf.set_auto_page_break(True, margin=15)   # type: ignore[union-attr]
            pdf.set_y(max(t_end, c_end) + 3.0)         # type: ignore[union-attr]

        def _section_1col(title: str, headers: list[str], rows: list[tuple],
                          col_widths: list[float]) -> None:
            """Sección con tabla a ancho completo (para Resumen)."""
            _check_space(60.0)
            _section_title(title)
            pdf.set_fill_color(210, 210, 210)    # type: ignore[union-attr]
            pdf.set_text_color(0, 0, 0)          # type: ignore[union-attr]
            pdf.set_font("Helvetica", "B", 10)   # type: ignore[union-attr]
            for h, w in zip(headers, col_widths):
                pdf.cell(w, HDR_H, h, border=1, fill=True)  # type: ignore[union-attr]
            pdf.ln()                             # type: ignore[union-attr]
            for i, row_data in enumerate(rows):
                if i % 2 == 0:
                    pdf.set_fill_color(242, 242, 242)   # type: ignore[union-attr]
                else:
                    pdf.set_fill_color(255, 255, 255)   # type: ignore[union-attr]
                pdf.set_text_color(30, 30, 30)   # type: ignore[union-attr]
                pdf.set_font("Helvetica", size=9)  # type: ignore[union-attr]
                for val, w in zip(row_data, col_widths):
                    pdf.cell(w, DATA_H, str(val)[:35], border="B", fill=True)  # type: ignore[union-attr]
                pdf.ln()                          # type: ignore[union-attr]
            pdf.ln(3)                             # type: ignore[union-attr]

        # ── Secciones ─────────────────────────────────────────────────── #
        hours = stats.total_duration_ms / 3_600_000
        _section_1col(
            "Resumen general", ["Indicador", "Valor"],
            [
                ("Total pistas",               stats.total_tracks),
                ("Total artistas",             stats.total_artists),
                ("Total albums",               stats.total_albums),
                ("Horas de musica",            f"{hours:.1f} h"),
                ("Nacionalidades de artistas", stats.total_countries),
            ],
            [100.0, 60.0],
        )

        if stats.genre_counts:
            _section_2col("Generos", ["Genero", "Pistas"],
                          sorted(stats.genre_counts.items(), key=lambda x: x[1], reverse=True))
        if stats.decade_counts:
            _section_2col("Decadas", ["Decada", "Pistas"],
                          sorted(stats.decade_counts.items()))
        if stats.format_counts:
            _section_2col("Formatos", ["Formato", "Pistas"],
                          sorted(stats.format_counts.items(), key=lambda x: x[1], reverse=True))
        if stats.top_artists:
            _section_2col("Top 10 artistas por pistas", ["Artista", "Pistas"], stats.top_artists)
        if stats.album_type_counts:
            _section_2col("Tipo de album", ["Tipo", "Albums"],
                          sorted(stats.album_type_counts.items(), key=lambda x: x[1], reverse=True))
        if stats.artist_country_counts:
            _section_2col("Pais de origen de artistas", ["Pais", "Artistas"],
                          sorted(stats.artist_country_counts.items(), key=lambda x: x[1], reverse=True))
        if stats.label_country_counts:
            _section_2col("Pais de origen de sellos", ["Pais", "Sellos"],
                          sorted(stats.label_country_counts.items(), key=lambda x: x[1], reverse=True))

    # ------------------------------------------------------------------
    # XLSX — combinado y por sección
    # ------------------------------------------------------------------

    def export_xlsx(self, tracks: list[Track], stats: LibraryStats, filepath: str) -> None:
        """Exporta a Excel con dos hojas: Biblioteca y Estadísticas."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Biblioteca"
        self._write_library_sheet(ws1, tracks)
        ws2 = wb.create_sheet("Estadísticas")
        self._write_stats_sheet(ws2, stats)
        wb.save(filepath)

    def export_library_xlsx(self, tracks: list[Track], filepath: str) -> None:
        """Exporta solo la biblioteca a Excel (una hoja)."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Biblioteca"
        self._write_library_sheet(ws, tracks)
        wb.save(filepath)

    def export_stats_xlsx(self, stats: LibraryStats, filepath: str) -> None:
        """Exporta solo las estadísticas a Excel (una hoja)."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estadísticas"
        self._write_stats_sheet(ws, stats)
        wb.save(filepath)

    # ------------------------------------------------------------------
    # PDF — combinado y por sección
    # ------------------------------------------------------------------

    def export_pdf(self, tracks: list[Track], stats: LibraryStats, filepath: str) -> None:
        """Exporta a PDF con dos secciones: Biblioteca y Estadísticas."""
        from fpdf import FPDF
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.set_margins(10, 10, 10)
        self._write_library_pdf(pdf, tracks, stats)
        self._write_stats_pdf(pdf, stats)
        pdf.output(filepath)

    def export_library_pdf(self, tracks: list[Track], stats: LibraryStats, filepath: str) -> None:
        """Exporta solo la biblioteca a PDF."""
        from fpdf import FPDF
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.set_margins(10, 10, 10)
        self._write_library_pdf(pdf, tracks, stats)
        pdf.output(filepath)

    def export_stats_pdf(self, stats: LibraryStats, filepath: str) -> None:
        """Exporta solo las estadísticas a PDF."""
        from fpdf import FPDF
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.set_margins(10, 10, 10)
        self._write_stats_pdf(pdf, stats)
        pdf.output(filepath)

    # ------------------------------------------------------------------
    # CSV — biblioteca y estadísticas
    # ------------------------------------------------------------------

    def export_csv(self, tracks: list[Track], filepath: str) -> None:
        """Exporta solo la biblioteca a CSV (UTF-8 con BOM para Excel)."""
        fieldnames = ["#", "Título", "Artista", "Álbum", "Año", "Género", "Duración", "Formato"]
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for i, t in enumerate(tracks, 1):
                writer.writerow({
                    "#":         i,
                    "Título":    t.title       or "",
                    "Artista":   t.artist_name or "",
                    "Álbum":     t.album_title or "",
                    "Año":       t.year        or "",
                    "Género":    t.genre       or "",
                    "Duración":  _ms_to_str(t.duration_ms),
                    "Formato":   t.format.value if t.format else "",
                })

    def export_stats_csv(self, stats: LibraryStats, filepath: str) -> None:
        """Exporta las estadísticas a CSV en formato Sección,Indicador,Valor."""
        hours = stats.total_duration_ms / 3_600_000
        rows: list[tuple[str, str, str]] = [
            ("Resumen", "Total pistas",               str(stats.total_tracks)),
            ("Resumen", "Total artistas",             str(stats.total_artists)),
            ("Resumen", "Total álbumes",              str(stats.total_albums)),
            ("Resumen", "Horas de música",            f"{hours:.1f}"),
            ("Resumen", "Nacionalidades de artistas", str(stats.total_countries)),
        ]
        for g, c in sorted(stats.genre_counts.items(), key=lambda x: x[1], reverse=True):
            rows.append(("Géneros", g, str(c)))
        for d, c in sorted(stats.decade_counts.items()):
            rows.append(("Décadas", d, str(c)))
        for f, c in sorted(stats.format_counts.items(), key=lambda x: x[1], reverse=True):
            rows.append(("Formatos", f, str(c)))
        for artist, count in stats.top_artists:
            rows.append(("Top artistas", artist, str(count)))
        for t, c in sorted(stats.album_type_counts.items(), key=lambda x: x[1], reverse=True):
            rows.append(("Tipo de álbum", t, str(c)))
        for country, count in sorted(stats.artist_country_counts.items(), key=lambda x: x[1], reverse=True):
            rows.append(("País de artistas", country, str(count)))
        for country, count in sorted(stats.label_country_counts.items(), key=lambda x: x[1], reverse=True):
            rows.append(("País de sellos", country, str(count)))

        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["Sección", "Indicador", "Valor"])
            writer.writerows(rows)
